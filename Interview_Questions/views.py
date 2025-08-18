from rest_framework import generics, permissions
from .models import Quiz
from .serializers import QuizSerializer,QuizTitleSerializer
from .permissions import IsEmployee

# Create Quiz (already provided)
class QuizCreateView(generics.CreateAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    permission_classes = [IsEmployee]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

# Get All Quizzes
class QuizListView(generics.ListAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    permission_classes = [permissions.IsAuthenticated]

class MyQuizListView(generics.ListAPIView):
    serializer_class = QuizSerializer
    permission_classes = [IsEmployee]  # or IsAuthenticated if more general

    def get_queryset(self):
        return Quiz.objects.filter(created_by=self.request.user)


# Get Quiz by ID
class QuizDetailView(generics.RetrieveAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    permission_classes = [permissions.IsAuthenticated]

# Update/Edit Quiz (only employee who created can edit)
class QuizUpdateView(generics.UpdateAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    permission_classes = [IsEmployee]

    def get_queryset(self):
        return Quiz.objects.filter(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save()

    def partial_update(self, request, *args, **kwargs):

        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

# Delete Quiz (only employee who created can delete)
class QuizDeleteView(generics.DestroyAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    permission_classes = [IsEmployee]

    def get_queryset(self):
        return Quiz.objects.filter(created_by=self.request.user)
    

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Quiz, Question, Option  # Adjust import based on your models

class ExportQuizExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Get module filter from query params
        module = request.query_params.get('module', None)

        # Create the workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Interview Questions"

        # Write header
        headers = ["Quiz Title", "Question Text", "Option", "Is Correct"]
        ws.append(headers)

        # Gather and write data
        quizzes = Quiz.objects.prefetch_related('questions__options').all()
        if module:
            quizzes = quizzes.filter(title__icontains=module)

        if not quizzes.exists():
            return Response(
                {"detail": "No quizzes found for the specified module."},
                status=status.HTTP_404_NOT_FOUND
            )

        for quiz in quizzes:
            for question in quiz.questions.all():
                for option in question.options.all():
                    ws.append([
                        quiz.title,
                        question.text,
                        option.text,
                        "Yes" if option.is_correct else "No"
                    ])

        # Set column widths
        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 30

        # Prepare response
        filename = f"interview_questions_{module or 'all'}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response


class MyQuizTitlesView(generics.ListAPIView):
    serializer_class = QuizTitleSerializer
    permission_classes = [IsEmployee]

    def get_queryset(self):
        return Quiz.objects.filter(created_by=self.request.user)


# pip install reportlab Pillow

from django.http import HttpResponse, Http404
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from io import BytesIO
from datetime import datetime
import os
from django.conf import settings
from internships.models import Internship

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def download_quiz_pdf(request, quiz_id):
#     try:
#         quiz = Quiz.objects.get(id=quiz_id, created_by=request.user)
#         internship = Internship.objects.filter(quiz_set=quiz).first()
#         passing_score = f"{internship.pass_percentage}%" if internship and internship.pass_percentage else "N/A"
        
#         buffer = BytesIO()
#         doc = SimpleDocTemplate(
#             buffer, 
#             pagesize=A4, 
#             rightMargin=25*mm, 
#             leftMargin=25*mm, 
#             topMargin=10*mm,
#             bottomMargin=25*mm
#         )
        
#         # Get styles
#         styles = getSampleStyleSheet()
        
#         # Custom styles
#         title_style = ParagraphStyle(
#             'CustomTitle',
#             parent=styles['Heading1'],
#             fontSize=18,
#             spaceAfter=20,
#             spaceBefore=10,
#             alignment=TA_CENTER,
#             textColor=colors.HexColor('#1e40af'),
#             fontName='Helvetica-Bold'
#         )
        
#         subtitle_style = ParagraphStyle(
#             'SubtitleStyle',
#             parent=styles['Normal'],
#             fontSize=12,
#             spaceAfter=15,
#             alignment=TA_CENTER,
#             textColor=colors.HexColor('#64748b'),
#             fontName='Helvetica'
#         )
        
#         # Updated question header style - removed box styling
#         question_header_style = ParagraphStyle(
#             'QuestionHeader',
#             parent=styles['Heading2'],
#             fontSize=14,
#             spaceAfter=8,
#             spaceBefore=15,
#             leftIndent=0,
#             fontName='Helvetica-Bold',
#             textColor=colors.HexColor('#1e40af')
#         )
        
#         question_text_style = ParagraphStyle(
#             'QuestionText',
#             parent=styles['Normal'],
#             fontSize=11,
#             spaceAfter=10,
#             spaceBefore=5,
#             leftIndent=0,
#             fontName='Helvetica',
#             textColor=colors.HexColor('#374151')
#         )
        
#         option_style = ParagraphStyle(
#             'OptionStyle',
#             parent=styles['Normal'],
#             fontSize=10,
#             spaceAfter=4,
#             leftIndent=15,
#             fontName='Helvetica',
#             textColor=colors.HexColor('#4b5563')
#         )
        
#         correct_answer_style = ParagraphStyle(
#             'CorrectAnswerStyle',
#             parent=styles['Normal'],
#             fontSize=10,
#             spaceAfter=15,
#             spaceBefore=8,
#             leftIndent=15,
#             textColor=colors.HexColor('#16a34a'),
#             fontName='Helvetica-Bold',
#             backColor=colors.HexColor('#f0fdf4'),
#             borderWidth=1,
#             borderColor=colors.HexColor('#bbf7d0'),
#             borderPadding=5
#         )
        
#         # Build PDF content
#         story = []
        
#         # Add document border frame for first page
#         border_frame = Table([['']], colWidths=[160*mm], rowHeights=[250*mm])
#         border_frame.setStyle(TableStyle([
#             ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
#             ('VALIGN', (0, 0), (-1, -1), 'TOP'),
#         ]))
        
#         # Create first page content
#         first_page_content = []
        
#         # Company Logo and Header
#         try:
#             # Update this path to your actual logo path
#             logo_path = os.path.join(settings.MEDIA_ROOT, 'company_logo.png')
#             # Alternative: logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo.png')
            
#             if os.path.exists(logo_path):
#                 logo = Image(logo_path, width=60*mm, height=20*mm)
#                 logo.hAlign = 'CENTER'
#                 first_page_content.append(logo)
#                 first_page_content.append(Spacer(1, 20))
#             else:
#                 # Fallback company name if logo not found
#                 company_style = ParagraphStyle(
#                     'CompanyStyle',
#                     parent=styles['Normal'],
#                     fontSize=25,
#                     fontName='Helvetica-Bold',
#                     textColor=colors.HexColor('#1e40af'),
#                     alignment=TA_LEFT
#                 )
#                 first_page_content.append(Paragraph("Internship", company_style))
#                 first_page_content.append(Spacer(1, 55))
                
#         except Exception as e:
#             # Fallback if image processing fails
#             company_style = ParagraphStyle(
#                 'CompanyStyle',
#                 parent=styles['Normal'],
#                 fontSize=16,
#                 fontName='Helvetica-Bold',
#                 textColor=colors.HexColor('#1e40af'),
#                 alignment=TA_CENTER
#             )
#             first_page_content.append(Paragraph("Internship", company_style))
#             first_page_content.append(Spacer(1, 55))
        
#         # Quiz Title
#         first_page_content.append(Paragraph(f"<b>Quiz Title:</b> {quiz.title}", title_style))
#         first_page_content.append(Paragraph("Interview Quiz Assessment", subtitle_style))
#         first_page_content.append(Spacer(1, 20))
        
#         # Quiz Information Table
#         quiz_info_data = [
#             ['Quiz Details', ''],
#             ['Duration', f'{quiz.duration_minutes} minutes'],
#             ['Total Questions', str(quiz.questions.count())],
#             ['Passing Score', passing_score],
#             ['Generated On', datetime.now().strftime('%B %d, %Y at %I:%M %p')],
#             ['Quiz ID', str(quiz.id)]
#         ]
        
#         quiz_info_table = Table(quiz_info_data, colWidths=[60*mm, 80*mm], hAlign='CENTER')

#         quiz_info_table.setStyle(TableStyle([
#             # Header row
#             ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('FONTSIZE', (0, 0), (-1, 0), 12),
#             ('SPAN', (0, 0), (1, 0)),
#             ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
#             # Data rows
#             ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
#             ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
#             ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
#             ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
#             ('FONTSIZE', (0, 1), (-1, -1), 11),
#             ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            
#             # Borders
#             ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('TOPPADDING', (0, 0), (-1, -1), 10),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
#             ('LEFTPADDING', (0, 0), (-1, -1), 15),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 15),
#         ]))
#         first_page_content.append(Spacer(1, 20))
#         first_page_content.append(quiz_info_table)
#         first_page_content.append(Spacer(1, 50))
        
#         # Instructions
#         instructions_style = ParagraphStyle(
#             'InstructionsStyle',
#             parent=styles['Normal'],
#             fontSize=11,
#             spaceAfter=20,
#             fontName='Helvetica',
#             textColor=colors.HexColor('#6b7280'),
#             alignment=TA_JUSTIFY
#         )
#         first_page_content.append(Paragraph(
#              "<b>Instructions:</b><br/><br/>"
#     "• This document contains the complete question set for the interview quiz.<br/><br/>"
#     "• Each question is followed by multiple choice options.<br/><br/>"
#     "• The correct answer is clearly marked below each question.<br/><br/>"
#     "• Please review all questions thoroughly before conducting the assessment.<br/><br/>"
#     "• Ensure proper time management during the quiz administration.",
#             instructions_style
#         ))
        
#         # Add all first page content to story
#         story.extend(first_page_content)
        
#         # Add page break to start questions on second page
#         story.append(PageBreak())
        
#         # Questions Section Header for second page
#         questions_header = ParagraphStyle(
#             'QuestionsHeader',
#             parent=styles['Heading1'],
#             fontSize=18,
#             spaceAfter=20,
#             spaceBefore=10,
#             fontName='Helvetica-Bold',
#             textColor=colors.HexColor('#1e40af'),
#             alignment=TA_CENTER
#         )
#         story.append(Paragraph("Quiz Questions", questions_header))
        
       
#         story.append(Spacer(1, 20))
        
#         # Questions and Options
#         for index, question in enumerate(quiz.questions.all(), 1):
#             # Question Header - clean design without box
#             story.append(Paragraph(f"Question {index}:", question_header_style))
            
#             # Question Text
#             story.append(Paragraph(question.text, question_text_style))
#             story.append(Spacer(1, 10))
            
#             # Options
#             options = question.options.all()
#             correct_option = None
#             correct_letter = None
            
#             # Display options in a clean format
#             for opt_index, option in enumerate(options):
#                 option_letter = chr(65 + opt_index)  # A, B, C, D
                
#                 if option.is_correct:
#                     correct_option = option.text
#                     correct_letter = option_letter
                
#                 option_text = f"<b>{option_letter}.</b> {option.text}"
#                 story.append(Paragraph(option_text, option_style))
            
#             story.append(Spacer(1, 10))
            
#             # Correct Answer
#             if correct_option and correct_letter:
#                 story.append(Paragraph(
#                     f"✓ <b>Correct Answer:</b> Option {correct_letter} - {correct_option}", 
#                     correct_answer_style
#                 ))
            
#             story.append(Spacer(1, 25))
            
#             # Add page break after every 4 questions for better readability
#             if index % 4 == 0 and index < quiz.questions.count():
#                 story.append(PageBreak())
        
#         # Footer
#         footer_style = ParagraphStyle(
#             'FooterStyle',
#             parent=styles['Normal'],
#             fontSize=8,
#             alignment=TA_CENTER,
#             textColor=colors.HexColor('#9ca3af'),
#             spaceBefore=30
#         )
        
#         footer_table = Table([['']], colWidths=[160*mm])
#         footer_table.setStyle(TableStyle([
#             ('LINEABOVE', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
#         ]))
#         story.append(footer_table)
#         story.append(Spacer(1, 10))
#         story.append(Paragraph(
#             f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | "
#             f"Quiz ID: {quiz.id} | Confidential & Proprietary",
#             footer_style
#         ))
        
#         # Build PDF
#         def draw_page_border(canvas, doc):
#             canvas.saveState()
#             border_color = colors.HexColor('#1e40af')
#             canvas.setStrokeColor(border_color)  # black color
#             canvas.setLineWidth(1.5)             # thickness of border
#             width, height = doc.pagesize
#             margin = 10  # distance from page edges
#             canvas.rect(margin, margin, width - 2*margin, height - 2*margin)
#             canvas.restoreState()
#         doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)

        
#         # Get PDF data
#         pdf_data = buffer.getvalue()
#         buffer.close()
        
#         # Create response
#         safe_filename = "".join(c for c in quiz.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
#         filename = f"{safe_filename.replace(' ', '_').lower()}_quiz.pdf"
        
#         response = HttpResponse(pdf_data, content_type='application/pdf')
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         response['Content-Length'] = len(pdf_data)
        
#         return response
        
#     except Quiz.DoesNotExist:
#         raise Http404("Quiz not found or you don't have permission to access it")
#     except Exception as e:
#         return Response({'error': f'Failed to generate PDF: {str(e)}'}, status=500)







@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_quiz_pdf(request, quiz_id):
    try:
        quiz = Quiz.objects.get(id=quiz_id, created_by=request.user)
        internship = Internship.objects.filter(quiz_set=quiz).first()
        passing_score = f"{internship.pass_percentage}%" if internship and internship.pass_percentage else "N/A"

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            rightMargin=25*mm, 
            leftMargin=25*mm, 
            topMargin=10*mm,
            bottomMargin=25*mm
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            spaceBefore=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e40af'),
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=15,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b'),
            fontName='Helvetica'
        )

        question_text_style = ParagraphStyle(
            'QuestionText',
            parent=styles['Normal'],
            fontSize=14,  # increased font size
            spaceAfter=10,
            spaceBefore=5,
            leftIndent=0,
            fontName='Helvetica-Bold',  # Questions bold
            textColor=colors.HexColor('#374151')
        )

        option_style = ParagraphStyle(
            'OptionStyle',
            parent=styles['Normal'],
            fontSize=12,  # increased font size
            spaceAfter=2,  # 1 line spacing between options
            leftIndent=15,
            fontName='Helvetica',  # normal font
            textColor=colors.HexColor('#4b5563')
        )

        correct_answer_style = ParagraphStyle(
            'CorrectAnswerStyle',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=10,
            spaceBefore=5,
            leftIndent=15,
            textColor=colors.HexColor('#16a34a'),
            fontName='Helvetica-Bold',
            backColor=colors.HexColor('#f0fdf4'),
            borderWidth=1,
            borderColor=colors.HexColor('#34d399'),
            borderPadding=4  # slightly smaller padding for compact box
        )

        story = []

        # First page content remains unchanged
        border_frame = Table([['']], colWidths=[160*mm], rowHeights=[250*mm])
        border_frame.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        first_page_content = []

        try:
            logo_path = os.path.join(settings.MEDIA_ROOT, 'company_logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=60*mm, height=20*mm)
                logo.hAlign = 'CENTER'
                first_page_content.append(logo)
                first_page_content.append(Spacer(1, 20))
            else:
                company_style = ParagraphStyle(
                    'CompanyStyle',
                    parent=styles['Normal'],
                    fontSize=25,
                    fontName='Helvetica-Bold',
                    textColor=colors.HexColor('#1e40af'),
                    alignment=TA_LEFT
                )
                first_page_content.append(Paragraph("Internship", company_style))
                first_page_content.append(Spacer(1, 55))
        except Exception as e:
            company_style = ParagraphStyle(
                'CompanyStyle',
                parent=styles['Normal'],
                fontSize=16,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1e40af'),
                alignment=TA_CENTER
            )
            first_page_content.append(Paragraph("Internship", company_style))
            first_page_content.append(Spacer(1, 55))

        first_page_content.append(Paragraph(f"<b>Quiz Title:</b> {quiz.title}", title_style))
        first_page_content.append(Paragraph("Interview Quiz Assessment", subtitle_style))
        first_page_content.append(Spacer(1, 10))

        quiz_info_data = [
            ['Quiz Details', ''],
            ['Duration', f'{quiz.duration_minutes} minutes'],
            ['Total Questions', str(quiz.questions.count())],
            ['Passing Score', passing_score],
            ['Generated On', datetime.now().strftime('%B %d, %Y at %I:%M %p')],
            ['Quiz ID', str(quiz.id)]
        ]

        quiz_info_table = Table(quiz_info_data, colWidths=[60*mm, 80*mm], hAlign='CENTER')
        quiz_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('SPAN', (0, 0), (1, 0)),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ]))

        first_page_content.append(Spacer(1, 20))
        first_page_content.append(quiz_info_table)
        first_page_content.append(Spacer(1, 50))

        instructions_style = ParagraphStyle(
            'InstructionsStyle',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=20,
            fontName='Helvetica',
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_JUSTIFY
        )
        first_page_content.append(Paragraph(
             "<b>Instructions:</b><br/><br/>"
             "• This document contains the complete question set for the interview quiz.<br/><br/>"
             "• Each question is followed by multiple choice options.<br/><br/>"
             "• The correct answer is clearly marked below each question.<br/><br/>"
             "• Please review all questions thoroughly before conducting the assessment.<br/><br/>"
             "• Ensure proper time management during the quiz administration.",
             instructions_style
        ))

        story.extend(first_page_content)
        story.append(PageBreak())

        questions_header = ParagraphStyle(
            'QuestionsHeader',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            spaceBefore=10,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1e40af'),
            alignment=TA_CENTER
        )
        story.append(Paragraph("Quiz Questions", questions_header))
        story.append(Spacer(1, 20))

        # Questions and Options
        for index, question in enumerate(quiz.questions.all(), 1):
            # Question in bold
            story.append(Paragraph(f"Q{index}: {question.text}", question_text_style))
            story.append(Spacer(1, 15))

            # Options with 1 line space
            for opt_index, option in enumerate(question.options.all()):
                option_letter = chr(65 + opt_index)
                story.append(Paragraph(f"{option_letter}. {option.text}", option_style))
                story.append(Spacer(1, 8))  # 1 line spacing between options

            # Correct answer in compact boxed style
            correct_option = next((o for o in question.options.all() if o.is_correct), None)
            if correct_option:
                correct_table = Table([[Paragraph(f"✓ Correct Answer: {correct_option.text}", correct_answer_style)]],
                                      colWidths=[150*mm])
                correct_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#d1fae5')),
                    ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#34d399')),
                    ('LEFTPADDING', (0,0), (-1,-1), 6),
                    ('RIGHTPADDING', (0,0), (-1,-1), 6),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ]))
                story.append(correct_table)

            story.append(Spacer(1, 20))

        # Footer remains unchanged
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#9ca3af'),
            spaceBefore=30
        )
        footer_table = Table([['']], colWidths=[160*mm])
        footer_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        story.append(footer_table)
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | "
            f"Quiz ID: {quiz.id} | Confidential & Proprietary",
            footer_style
        ))

        def draw_page_border(canvas, doc):
            canvas.saveState()
            border_color = colors.HexColor('#1e40af')
            canvas.setStrokeColor(border_color)
            canvas.setLineWidth(1.5)
            width, height = doc.pagesize
            margin = 10
            canvas.rect(margin, margin, width - 2*margin, height - 2*margin)
            canvas.restoreState()

        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)

        pdf_data = buffer.getvalue()
        buffer.close()

        safe_filename = "".join(c for c in quiz.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"{safe_filename.replace(' ', '_').lower()}_quiz.pdf"

        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(pdf_data)

        return response

    except Quiz.DoesNotExist:
        raise Http404("Quiz not found or you don't have permission to access it")
    except Exception as e:
        return Response({'error': f'Failed to generate PDF: {str(e)}'}, status=500)